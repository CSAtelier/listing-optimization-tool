from amazoncaptcha import AmazonCaptcha
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import bs4
import time
import re
from openpyxl import Workbook
import random

def asin_to_url(asin_list):
    url_list = []
    for asin in asin_list:
        url_list.append(f"https://www.amazon.com/dp/{asin}")
        url_list.append(f"https://www.amazon.ca/dp/{asin}")
    return url_list

def captcha_handle(response,driver):
    img = response.find_all("img")[0]["src"]
    captcha = AmazonCaptcha.fromlink(img)
    solution = captcha.solve()
    driver.find_element(By.NAME, "field-keywords").send_keys(solution) 
    driver.find_element(By.CLASS_NAME, "a-button-text").click()

def change_location_us(driver):
    
    driver.find_element(By.ID, "nav-global-location-popover-link").click()
    element = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "GLUXZipUpdateInput")))
    postcode_form = driver.find_element(By.ID, "GLUXZipUpdateInput").send_keys("73001") 
    postcode_button = driver.find_element(By.XPATH, '//*[@id="GLUXZipUpdate"]/span/input').click()
    time.sleep(1)
    continue_button = driver.find_element(By.XPATH, '//*[@id="a-popover-2"]/div/div[2]/span')
    continue_button.click()
    time.sleep(1)


def get_price_us(response):

    price = response.find('span', attrs = {'class':'a-offscreen'})
    price = re.search(r'\$\d+\.\d+', price.text).group()
    return price

def change_location_ca(driver):
    
    driver.find_element(By.ID, "nav-global-location-popover-link").click()
    element = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "GLUXZipUpdateInput_0")))
    postcode_form0 = driver.find_element(By.ID, "GLUXZipUpdateInput_0").send_keys("M5V") 
    postcode_form1 = driver.find_element(By.ID, "GLUXZipUpdateInput_1").send_keys("3L9") 
    postcode_button = driver.find_element(By.XPATH, '//*[@id="GLUXZipUpdate"]/span/input').click()
    time.sleep(1)
    # continue_button = driver.find_element(By.XPATH, '//*[@id="a-popover-3"]/div/div[2]/span').click()

def get_price_ca(response):

    price = response.find('span', attrs = {'class':'a-offscreen'})
    price = re.search(r'\$\d+\.\d+', price.text).group()
    return price

def create_excel(asin, price, country, index):

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'ASIN'
    ws['B1'] = 'CA Prices'
    ws['C1'] = 'US Prices'
    ws['D1'] = 'Revenue Difference'
    ws['E1'] = 'URL Amazon US'
    ws['F1'] = 'URL Amazon CA'
    ws[f'A{index+2}'] = asin
    if country == 'ca':
        ws[f'B{index+2}'] = price
    elif country == 'us':
        ws[f'C{index+2}'] = price
    return wb
    
def extract_asin(url):
    parts = url.split('/')
    dp_index = parts.index('dp')
    asin = parts[dp_index + 1]
    return asin

    