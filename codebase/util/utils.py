from openpyxl import Workbook
import random
from codebase.dataset_loader import DatasetLoader


def asin_to_url(asin_list):
    url_list_us = []
    url_list_ca = []
    for asin in asin_list:
        url_list_us.append(f"https://www.amazon.com/dp/{asin}")
        url_list_ca.append(f"https://www.amazon.ca/dp/{asin}")
    return url_list_us, url_list_ca

def create_excel(price_dict_us, price_dict_ca, save_path):

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'ASIN'
    ws['B1'] = 'US Prices'
    ws['C1'] = 'US Sale'
    ws['D1'] = 'CA Prices'
    ws['E1'] = 'CA Sale'
    ws['F1'] = 'US/CA'
    ws['G1'] = 'URL Amazon US'
    ws['H1'] = 'URL Amazon CA'
    key_list = list(price_dict_us.keys())
    for i in range(len(key_list)):
        ws[f'A{i+2}'] = key_list[i]
        ws[f'B{i+2}'] = price_dict_us[key_list[i]][0]
        ws[f'C{i+2}'] = float(price_dict_us[key_list[i]][1].replace(',', '.'))
        ws[f'D{i+2}'] = price_dict_ca[key_list[i]][0]
        ws[f'E{i+2}'] = float(price_dict_ca[key_list[i]][1].replace(',', '.'))
        ws[f'G{i+2}'] = f"https://www.amazon.com/dp/{key_list[i]}"
        ws[f'H{i+2}'] = f"https://www.amazon.ca/dp/{key_list[i]}"
        if price_dict_ca[key_list[i]] == 0 or price_dict_us[key_list[i]] == 0:
            ws[f'F{i+2}'] = 0
        else:
            ws[f'F{i+2}'] = price_dict_us[key_list[i]][0]/price_dict_ca[key_list[i]][0]
    wb.save(save_path + '/us-ca_excel_'+str(random.randint(1,10000))+'.xlsx')
    return wb
    

def extract_asin(url):
    parts = url.split('/')
    dp_index = parts.index('dp')
    asin = parts[dp_index + 1]
    return asin

    