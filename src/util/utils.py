import openpyxl
import random
from src.dataset_loader import DatasetLoader
from currency_converter import CurrencyConverter
import os

def asin_to_url(asin_list):
    url_list_us = []
    url_list_ca = []
    for asin in asin_list:
        url_list_us.append(f"https://www.amazon.com/dp/{asin}")
        url_list_ca.append(f"https://www.amazon.ca/dp/{asin}")
    return url_list_us, url_list_ca

def create_excel(price_dict_us, price_dict_ca, data_path,
                 us_price_column=None, ca_price_column=None,
                 us_sale_column=None, ca_sale_column=None,
                 revenue_column=None, excel_index=0):

    # Set the path for the Excel file
    excel_path = data_path.replace('.csv', '.xlsx')
    print(excel_path)

    # Check if the Excel file already exists, if not, convert CSV to Excel
    if not os.path.exists(excel_path):
        # Convert CSV to Excel
        csv_data = pd.read_csv(data_path)
        csv_data.to_excel(excel_path, index=False)
    
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Process and populate data
    key_list = list(price_dict_us.keys())
    for i in range(len(key_list)):
        if us_price_column:
            ws[us_price_column + f'{excel_index+2}'] = price_dict_us[key_list[i]][0]
        if ca_price_column:
            ws[ca_price_column + f'{excel_index+2}'] = price_dict_ca[key_list[i]][0]
        if us_sale_column:
            ws[us_sale_column + f'{excel_index+2}'] = float(price_dict_us[key_list[i]][1])
        if ca_sale_column:
            ws[ca_sale_column + f'{excel_index+2}'] = float(price_dict_ca[key_list[i]][1])
        if revenue_column:
            c = CurrencyConverter()
            usd_cad = c.convert(1, 'USD', 'CAD')
            usd_cad_price = (float(price_dict_us[key_list[i]][0]) * 1.06) * usd_cad
            shipping = 2.5 * usd_cad
            cost = usd_cad_price + shipping
            print(usd_cad_price, shipping, cost, float(price_dict_ca[key_list[i]][2]))
            ws[revenue_column + f'{excel_index+2}'] = ((float(price_dict_ca[key_list[i]][2]) - cost) * 100.0) / cost

        # Calculate the price ratio if both prices are available
        if price_dict_ca[key_list[i]] == 0 or price_dict_us[key_list[i]] == 0:
            ws[f'F{excel_index+2}'] = 0
        else:
            try:
                ws[f'F{excel_index+2}'] = price_dict_us[key_list[i]][0] / price_dict_ca[key_list[i]][0]
            except:
                ws[f'F{excel_index+2}'] = 0

    # Save the updated workbook
    wb.save(excel_path)

import pandas as pd

def create_csv(price_dict_us, price_dict_ca,
               data_path, us_price_column=None, ca_price_column=None,
               us_sale_column=None, ca_sale_column=None,
               revenue_column=None, excel_index=0):
    
    # Read the existing CSV file into a DataFrame
    df = pd.read_csv(data_path)
    
    # Prepare to update the DataFrame
    key_list = list(price_dict_us.keys())
    
    for i in range(len(key_list)):
        asin = key_list[i]
        
        # Create a new row if it doesn't exist
        if excel_index + 2 > len(df):  # Adjust for header
            df.loc[excel_index + 1] = [None] * len(df.columns)

        if us_price_column is not None:
            df.at[excel_index + 1, us_price_column] = price_dict_us[asin][0]
        if ca_price_column is not None:
            df.at[excel_index + 1, ca_price_column] = price_dict_ca[asin][0]
        if us_sale_column is not None:
            df.at[excel_index + 1, us_sale_column] = float(price_dict_us[asin][1])
        if ca_sale_column is not None:
            df.at[excel_index + 1, ca_sale_column] = float(price_dict_ca[asin][1])
        if revenue_column is not None:
            c = CurrencyConverter()
            usd_cad = c.convert(1, 'USD', 'CAD') 
            usd_cad_price = (float(price_dict_us[asin][0]) * 1.06) * usd_cad
            shipping = 2.5 * usd_cad
            cost = usd_cad_price + shipping
            print(usd_cad_price, shipping, cost, float(price_dict_ca[asin][2]))
            df.at[excel_index + 1, revenue_column] = ((float(price_dict_ca[asin][2]) - cost) * 100.0) / cost

        # Calculate the price ratio
        if price_dict_ca[asin] == 0 or price_dict_us[asin] == 0:
            df.at[excel_index + 1, 'Price_Ratio'] = 0  # Ensure you have a column named 'Price_Ratio' in your DataFrame
        else:
            try:
                df.at[excel_index + 1, 'Price_Ratio'] = price_dict_us[asin][0] / price_dict_ca[asin][0]
            except:
                df.at[excel_index + 1, 'Price_Ratio'] = 0

    # Save the updated DataFrame back to the CSV file
    df.to_csv(data_path, index=False)


    

def extract_asin(url):
    parts = url.split('/')
    dp_index = parts.index('dp')
    asin = parts[dp_index + 1]
    return asin

    