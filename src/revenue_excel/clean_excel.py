import openpyxl
# from src.parser.parse import *
import os
import csv
import pandas as pd 
from api.revenue_calculator_api import *
from src.revenue_calculator_api.entities import Country, Currency
from src.revenue_calculator_api.revenue_calculator import get_arbitrage_product

def revenue_calculator(data_path,column):

    with open(data_path, 'r', newline='') as csvfile:
        if data_path.endswith('xlsx'):
            df =  pd.DataFrame(pd.read_excel(data_path)) 
        else:
            df = pd.read_csv(data_path)

    for i in range(len(df['ASIN'])):
        try:
            product = get_arbitrage_product(df['ASIN'][i], base_country=Country.USA, target_country=Country.CANADA, base_currency=Currency.USD, target_currency=Currency.CAD, exchange_rate= 1.36 , cost_of_shipment = 2.5)
            print(product)
        # get_inf(asin=df['ASIN'][i])
        except:
            print('error')


    

    # options = Options()
    # # options.add_argument("--window-size=1920,1080")
    # options.add_argument("--window-size=1500,900")
    # service = Service(ChromeDriverManager().install())
    # driver = webdriver.Chrome(options=options)
    # wb = openpyxl.load_workbook(data_path)
    # ws = wb.active
    # ws[column+'1'] = 'Revenue'
    # driver.get('https://www.google.com/search?q=us+to+ca&rlz=1C5CHFA_enTR987TR987&oq=us+to+ca&gs_lcrp=EgZjaHJvbWUyBggAEEUYOTIHCAEQABiPAjIHCAIQABiPAjIHCAMQABiPAjIGCAQQRRg9MgYIBRBFGDwyBggGEEUYPdIBCDMyMDVqMGo3qAIAsAIA&sourceid=chrome&ie=UTF-8')
    # time.sleep(2)
    # html = driver.page_source
    # response = BeautifulSoup(html,features="lxml")
    # ca_usd = response.find('span', attrs = {'class':'DFlfde SwHCTb'})
    # ca_usd = ca_usd.text.replace(',', '.')
    # with open(data_path, 'r', newline='') as csvfile:
    #     if data_path.endswith('xlsx'):
    #         df =  pd.DataFrame(pd.read_excel(data_path)) 
    #     else:
    #         df = pd.read_csv(data_path)
    # for i in range(len(df['ASIN'])):
    #     try: 
    #         price_ca = parse_ratio(driver=driver, asin=df['ASIN'][i])
    #     except:
    #         price_ca = 1
    #     calc_us = (float(df['Price'][i])*1.06)+2.56
    #     calc_us = float(calc_us)*float(ca_usd)
    #     try:
    #         print(price_ca,calc_us,((float(price_ca) - (float(calc_us)))/float(calc_us)),column+f'{i+2}')
    #         print((price_ca - (calc_us)))
    #         ws[column+f'{i+2}'] = ((float(price_ca) - (float(calc_us)))/float(calc_us))
    #     except: 
    #         ws[column+f'{i+2}'] = 0
    # wb.save(data_path)
    # ws['A1'] = 'ASIN'
    # ws['B1'] = 'US Prices'
    # ws['C1'] = 'CA Prices'
    # ws['D1'] = 'Ratio'
    # ws['E1'] = 'URL Amazon US'
    # ws['F1'] = 'URL Amazon CA'
    # asin_values = []
    # us_prices = []
    # ca_prices = []
    # driver.get('https://www.google.com/search?q=us+to+ca&rlz=1C5CHFA_enTR987TR987&oq=us+to+ca&gs_lcrp=EgZjaHJvbWUyBggAEEUYOTIHCAEQABiPAjIHCAIQABiPAjIHCAMQABiPAjIGCAQQRRg9MgYIBRBFGDwyBggGEEUYPdIBCDMyMDVqMGo3qAIAsAIA&sourceid=chrome&ie=UTF-8')
    # time.sleep(2)
    # html = driver.page_source
    # response = BeautifulSoup(html,features="lxml")
    # ca_usd = response.find('span', attrs = {'class':'DFlfde SwHCTb'})
    # ca_usd = ca_usd.text.replace(',', '.')
    # with open(data_path, 'r', newline='') as csvfile:
    #     if data_path.endswith('xlsx'):
    #        df =  pd.DataFrame(pd.read_excel(data_path)) 
    #     else:
    #         df = pd.read_csv(data_path)
    # for i in range(len(df['ASIN'])):
    #     ws[f'A{i+2}'] = df['ASIN'][i]
    #     ws[f'B{i+2}'] = df['US Prices'][i]
    #     price_ca = parse_ratio(driver=driver, asin=df['ASIN'][i])
    #     ws[f'C{i+2}'] = price_ca
    #     calc_us = float(df['US Prices'][i])+4.0
    #     calc_ca = float(calc_us)*float(ca_usd)
    #     ws[f'D{i+2}'] = (price_ca - (calc_ca))/calc_ca
    #     # except:
    #     #     ws[f'D{i+2}'] = 0
            
    #     ws[f'E{i+2}'] = f"https://www.amazon.com/dp/{df['ASIN'][i]}"
    #     ws[f'F{i+2}'] = f"https://www.amazon.ca/dp/{df['ASIN'][i]}"
    #     # print(float(df['US Prices'][i])+4.0)
    # #     ws[f'C{i+2}'] = price_dict_ca[key_list[i]]
    # #     ws[f'E{i+2}'] = f"https://www.amazon.com/dp/{key_list[i]}"
    # #     ws[f'F{i+2}'] = f"https://www.amazon.ca/dp/{key_list[i]}"
    # #     if price_dict_ca[key_list[i]] == 0 or price_dict_us[key_list[i]] == 0:
    # #         ws[f'D{i+2}'] = 0
    # #     else:
    # #         ws[f'D{i+2}'] = price_dict_us[key_list[i]]/price_dict_ca[key_list[i]]
    # wb.save(save_path + '/us-ca_excel_'+str(random.randint(1,10000))+'.xlsx')
    # return wb
